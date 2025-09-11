#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import sys
import re
import pandas as pd
import openpyxl
from collections import defaultdict

class RIMAnalyzer:
    def __init__(self):
        self.ru_months = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 
                          'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    
    def analyze_file(self, filepath):
        """Анализ файла журнала событий"""
        try:
            # Структура для хранения событий
            events_data = {
                'overvoltage': {'A': [], 'B': [], 'C': []},
                'undervoltage': {'A': [], 'B': [], 'C': []}
            }
            
            # Читаем файл через openpyxl для обработки объединенных ячеек
            try:
                from openpyxl import load_workbook
                
                # Загружаем workbook
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                
                # Собираем все данные, пропуская объединенные ячейки
                rows = []
                skip_first = False
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    # Проверяем первую строку
                    if row_idx == 0 and row[0] and 'Журнал событий' in str(row[0]):
                        skip_first = True
                        continue
                    
                    # Преобразуем в список, заменяя None на пустые строки
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    rows.append(row_data)
                
                wb.close()
                
            except Exception as e:
                # Если openpyxl не справился, пробуем pandas
                try:
                    # Читаем через pandas с разными движками
                    try:
                        df = pd.read_excel(filepath, engine='openpyxl', header=None)
                    except:
                        df = pd.read_excel(filepath, header=None)
                    
                    # Проверяем первую строку
                    if not df.empty and df.iloc[0, 0] is not None and 'Журнал событий' in str(df.iloc[0, 0]):
                        df = df.iloc[1:]
                    
                    # Преобразуем в список строк
                    rows = []
                    for _, row in df.iterrows():
                        row_data = [str(val) if pd.notna(val) else "" for val in row]
                        rows.append(row_data)
                        
                except Exception as e2:
                    return {
                        'success': False,
                        'error': f"Не удалось прочитать файл: {str(e2)}",
                        'has_errors': False
                    }
            
            # Находим начало данных
            start_row = 0
            for idx, row in enumerate(rows):
                if len(row) < 2:
                    continue
                    
                cell0 = row[0]
                
                # Нашли заголовки
                if 'Время' in cell0 or 'время' in cell0.lower():
                    start_row = idx + 1
                    break
                    
                # Нашли первую дату
                if re.match(r'\d{2}\.\d{2}\.\d{4}', cell0):
                    start_row = idx
                    break
            
            # Обрабатываем данные
            for idx in range(start_row, len(rows)):
                try:
                    row = rows[idx]
                    if len(row) < 5:
                        continue
                    
                    # Читаем значения
                    datetime_str = row[0].strip()
                    event = row[1].strip()
                    
                    if not datetime_str or not event:
                        continue
                    
                    # Числовые значения
                    try:
                        voltage_str = row[2].replace(',', '.').strip()
                        percent_str = row[3].replace(',', '.').strip()
                        duration_str = row[4].replace(',', '.').strip()
                        
                        # Убираем возможные единицы измерения
                        voltage_str = re.sub(r'[^\d\.,\-]', '', voltage_str)
                        percent_str = re.sub(r'[^\d\.,\-]', '', percent_str)
                        duration_str = re.sub(r'[^\d\.,\-]', '', duration_str)
                        
                        voltage = float(voltage_str)
                        percent = float(percent_str)
                        duration = float(duration_str)
                    except:
                        continue
                    
                    # Фильтры
                    if duration <= 60:
                        continue
                    if abs(voltage - 11.50) < 0.001 or voltage == 0:
                        continue
                    
                    # Извлекаем месяц
                    date_match = re.match(r'(\d{2})\.(\d{2})\.(\d{4})', datetime_str)
                    if not date_match:
                        continue
                    month = int(date_match.group(2))
                    
                    # Определяем фазу и тип
                    phase = None
                    event_type = None
                    
                    event_lower = event.lower()
                    
                    # Ищем фазу
                    if 'фаза a' in event_lower or 'phase a' in event_lower:
                        phase = 'A'
                    elif 'фаза b' in event_lower or 'phase b' in event_lower:
                        phase = 'B'
                    elif 'фаза c' in event_lower or 'phase c' in event_lower:
                        phase = 'C'
                    
                    # События окончания
                    if phase and ('окончание' in event_lower or 'конец' in event_lower):
                        if 'провал' in event_lower or 'понижен' in event_lower:
                            event_type = 'undervoltage'
                        elif 'перенапряжение' in event_lower or 'повышен' in event_lower:
                            event_type = 'overvoltage'
                    
                    # Добавляем событие
                    if phase and event_type:
                        events_data[event_type][phase].append({
                            'voltage': voltage,
                            'month': month,
                            'duration': duration
                        })
                        
                except Exception as e:
                    continue
            
            return self._generate_result(events_data)
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Ошибка анализа: {str(e)}",
                'has_errors': False
            }
    
    def _generate_result(self, events_data):
        """Генерация результата анализа"""
        summary_parts = []
        has_errors = False
        details = {
            'overvoltage': {},
            'undervoltage': {}
        }
        
        # Обработка перенапряжений
        for phase in ['A', 'B', 'C']:
            events = events_data['overvoltage'][phase]
            if len(events) > 10:
                has_errors = True
                months = [e['month'] for e in events]
                min_month = min(months)
                max_month = max(months)
                
                if min_month == max_month:
                    period = self.ru_months[min_month-1]
                else:
                    period = f"{self.ru_months[min_month-1]}-{self.ru_months[max_month-1]}"
                
                voltages = [e['voltage'] for e in events]
                max_voltage = max(voltages)
                min_voltage_in_overvoltage = min(voltages)
                count = len(events)
                
                min_percent = ((min_voltage_in_overvoltage - 220) / 220) * 100
                max_percent = ((max_voltage - 220) / 220) * 100
                
                summary_parts.append(
                    f"Фаза {phase}: Перенапряжение {min_percent:.1f}-{max_percent:.1f}% (max {max_voltage:.0f}В) ({period}) - {count} событий"
                )
                
                details['overvoltage'][f'phase_{phase}'] = {
                    'count': count,
                    'max': max_voltage,
                    'period': period
                }
        
        # Обработка провалов
        for phase in ['A', 'B', 'C']:
            events = events_data['undervoltage'][phase]
            if len(events) > 10:
                has_errors = True
                months = [e['month'] for e in events]
                min_month = min(months)
                max_month = max(months)
                
                if min_month == max_month:
                    period = self.ru_months[min_month-1]
                else:
                    period = f"{self.ru_months[min_month-1]}-{self.ru_months[max_month-1]}"
                
                voltages = [e['voltage'] for e in events]
                min_voltage = min(voltages)
                max_voltage_in_undervoltage = max(voltages)
                count = len(events)
                
                min_percent = ((220 - max_voltage_in_undervoltage) / 220) * 100
                max_percent = ((220 - min_voltage) / 220) * 100
                
                summary_parts.append(
                    f"Фаза {phase}: Провал {min_percent:.1f}-{max_percent:.1f}% (min {min_voltage:.0f}В) ({period}) - {count} событий"
                )
                
                details['undervoltage'][f'phase_{phase}'] = {
                    'count': count,
                    'min': min_voltage,
                    'period': period
                }
        
        if not has_errors:
            total_events = sum(len(events) for events in events_data['overvoltage'].values())
            total_events += sum(len(events) for events in events_data['undervoltage'].values())
            
            if total_events > 0:
                summary = f"Обнаружено событий: {total_events}, но все менее 10 по каждому типу"
            else:
                summary = "Напряжение в пределах ГОСТ"
        else:
            summary = '; '.join(summary_parts)
        
        return {
            'success': True,
            'summary': summary,
            'has_errors': has_errors,
            'details': details
        }

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': 'No file path provided'}))
        sys.exit(1)
    
    try:
        analyzer = RIMAnalyzer()
        result = analyzer.analyze_file(sys.argv[1])
        print(json.dumps(result, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}))
